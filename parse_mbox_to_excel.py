#!/usr/bin/env python3
"""
Parse Daily Coding Problems mbox file and create Excel spreadsheet.
Run this script from the Daily Coding Problems directory.
"""

import mailbox
import re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import email.utils
import os

# Category keywords for auto-classification
DATA_STRUCTURES = {
    "Array": ["array", "list", "subarray", "contiguous", "elements", "index", "sorted array"],
    "Tree": ["tree", "binary tree", "bst", "node", "root", "leaf", "subtree", "inorder", "preorder", "postorder"],
    "Graph": ["graph", "vertex", "vertices", "edge", "edges", "directed", "undirected", "cycle", "connected", "adjacency"],
    "Stack": ["stack", "push", "pop", "lifo", "parentheses", "balanced", "brackets"],
    "Queue": ["queue", "dequeue", "enqueue", "fifo", "circular", "quack"],
    "Linked List": ["linked list", "pointer", "next", "head", "tail", "singly", "doubly"],
    "Hash Map": ["hash", "dictionary", "map", "lookup", "frequency"],
    "Heap": ["heap", "priority queue", "min-heap", "max-heap", "heapify"],
    "Trie": ["trie", "prefix", "autocomplete", "word search"],
    "Matrix": ["matrix", "grid", "2d", "row", "column", "cell", "board", "n x n"],
}

ALGORITHM_TYPES = {
    "Dynamic Programming": ["dynamic programming", "dp", "memoization", "tabulation", "optimal substructure"],
    "Recursion": ["recursive", "recursion", "backtracking", "base case"],
    "BFS/DFS": ["bfs", "dfs", "breadth-first", "depth-first", "traversal", "level order"],
    "Greedy": ["greedy", "optimal", "interval", "scheduling"],
    "Sorting": ["sort", "sorted", "order", "merge sort", "quick sort"],
    "Binary Search": ["binary search", "sorted", "mid", "logarithmic"],
    "Sliding Window": ["sliding window", "window", "substring", "k elements"],
    "Two Pointers": ["two pointers", "left", "right", "opposite ends"],
    "Bit Manipulation": ["bit", "bitwise", "xor", "binary", "mask", "shift"],
}


def categorize(text: str, categories: dict) -> list:
    """Find matching categories based on keywords."""
    text_lower = text.lower()
    matches = []
    for category, keywords in categories.items():
        for keyword in keywords:
            if keyword.lower() in text_lower:
                if category not in matches:
                    matches.append(category)
                break
    return matches

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MBOX_PATH = os.path.join(SCRIPT_DIR, "Emails/Import/DCP.mbox/mbox")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "daily_coding_problems_categorized.xlsx")

# Headers for spreadsheet (matching categorized format)
HEADERS = ['Problem #', 'Difficulty', 'Company', 'Title', 'Description',
           'DS Primary', 'DS Subtypes', 'Algo Primary', 'Algo Subtypes',
           'Status', 'Date Attempted', 'Date Solved', 'Email Date']


def load_existing_problems():
    """Load existing problems from Excel spreadsheet if it exists."""
    if not os.path.exists(OUTPUT_PATH):
        return {}

    print(f"Loading existing spreadsheet: {OUTPUT_PATH}")
    wb = load_workbook(OUTPUT_PATH)
    ws = wb.active

    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # Problem # column
            problem_num = str(row[0])
            existing[problem_num] = {
                'Problem #': problem_num,
                'Difficulty': row[1] or '',
                'Company': row[2] or '',
                'Title': row[3] or '',
                'Description': row[4] or '',
                'DS Primary': row[5] or '',
                'DS Subtypes': row[6] or '',
                'Algo Primary': row[7] or '',
                'Algo Subtypes': row[8] or '',
                'Status': row[9] or 'Not Started',
                'Date Attempted': row[10],
                'Date Solved': row[11],
                'Email Date': row[12] if len(row) > 12 else None
            }

    print(f"  Found {len(existing)} existing problems")
    wb.close()
    return existing


def parse_mbox(existing_problems=None):
    """Parse the mbox file and extract problem data, skipping existing ones."""
    if existing_problems is None:
        existing_problems = {}

    print("Parsing mbox file...")
    print(f"Reading from: {MBOX_PATH}")

    mbox = mailbox.mbox(MBOX_PATH)
    problems = []
    skipped = 0

    for idx, message in enumerate(mbox):
        if idx % 100 == 0:
            print(f"  Processed {idx} messages...")

        try:
            # Extract subject line
            subject = message.get('Subject', '')

            # Extract problem number and difficulty from subject
            # Format: "Daily Coding Problem: Problem #1116 [Hard]"
            match = re.search(r'Problem #(\d+)\s+\[(Easy|Medium|Hard)\]', subject)
            if not match:
                continue

            problem_num = match.group(1)
            difficulty = match.group(2)

            # Skip if problem already exists in spreadsheet
            if problem_num in existing_problems:
                skipped += 1
                continue
            
            # Extract email date
            date_str = message.get('Date', '')
            if date_str:
                parsed_date = email.utils.parsedate_to_datetime(date_str)
                email_date = parsed_date.strftime('%Y-%m-%d')
            else:
                email_date = ''
            
            # Get the email body
            body = ''
            if message.is_multipart():
                for part in message.walk():
                    if part.get_content_type() == 'text/plain':
                        body = part.get_payload(decode=True).decode('utf-8', errors='ignore')
                        break
            else:
                body = message.get_payload(decode=True).decode('utf-8', errors='ignore')
            
            # Extract company
            company_match = re.search(r'This problem was asked by ([^.]+)\.', body)
            company = company_match.group(1).strip() if company_match else ''
            
            # Extract problem description (everything after company line until the separator)
            desc_match = re.search(r'This problem was asked by [^.]+\.\s*\n\s*\n(.+?)(?:\n\s*-{50,}|\n\s*Upgrade to premium)', body, re.DOTALL)
            if desc_match:
                description = desc_match.group(1).strip()
                # Clean up the description - remove quoted-printable artifacts
                description = re.sub(r'=\s*\n', '', description)  # Remove quoted-printable line breaks
                description = re.sub(r'\s+', ' ', description)  # Normalize whitespace
                description = description[:1000]  # Reasonable limit for Excel cell
            else:
                description = ''
            
            # Extract title from first line or sentence of description
            title_match = re.match(r'^([^.!?]+[.!?])', description)
            title = title_match.group(1).strip() if title_match else ''
            if len(title) > 100:
                title = title[:100] + '...'

            # Skip problems with no description (empty content)
            if not description:
                continue

            # Use 'Unknown' if no company found
            if not company:
                company = 'Unknown'

            # Auto-categorize based on description keywords
            data_structures = categorize(description, DATA_STRUCTURES)
            algorithm_types = categorize(description, ALGORITHM_TYPES)

            # Split into primary (first match) and subtypes (rest)
            ds_primary = data_structures[0] if data_structures else ''
            ds_subtypes = ', '.join(data_structures[1:]) if len(data_structures) > 1 else ''
            algo_primary = algorithm_types[0] if algorithm_types else ''
            algo_subtypes = ', '.join(algorithm_types[1:]) if len(algorithm_types) > 1 else ''

            problems.append({
                'Problem #': problem_num,
                'Difficulty': difficulty,
                'Company': company,
                'Title': title,
                'Description': description,
                'DS Primary': ds_primary,
                'DS Subtypes': ds_subtypes,
                'Algo Primary': algo_primary,
                'Algo Subtypes': algo_subtypes,
                'Status': 'Not Started',
                'Date Attempted': '',
                'Date Solved': '',
                'Email Date': email_date
            })
            
        except Exception as e:
            print(f"  Error processing message {idx}: {e}")
            continue

    print(f"\nFound {len(problems)} new problems (skipped {skipped} existing)")
    return problems, skipped


def create_excel(new_problems, existing_problems=None):
    """Create or update Excel workbook with problem data."""
    if existing_problems is None:
        existing_problems = {}

    print("\nCreating Excel spreadsheet...")

    # Merge existing and new problems
    all_problems = list(existing_problems.values()) + new_problems

    # Sort by problem number
    all_problems.sort(key=lambda x: int(x['Problem #']))

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Coding Problems"

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data with proper typing
    for row_idx, problem in enumerate(all_problems, 2):
        for col_idx, header in enumerate(HEADERS, 1):
            value = problem[header]

            # Convert Problem # to integer
            if header == 'Problem #':
                value = int(value) if value else None

            # Handle date fields (may be string, datetime, or None)
            elif header in ['Date Attempted', 'Date Solved', 'Email Date']:
                if value is None:
                    pass  # Keep as None
                elif isinstance(value, datetime):
                    pass  # Already a datetime, keep it
                elif isinstance(value, str) and value.strip():
                    try:
                        value = datetime.strptime(value.strip(), '%Y-%m-%d')
                    except ValueError:
                        value = None  # Invalid date format
                else:
                    value = None

            # Keep strings as strings (Difficulty, Company, Title, Description, etc.)
            # Empty strings should remain as empty strings for text fields

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Wrap text for description column
            if header == 'Description':
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Format date columns
            if header in ['Date Attempted', 'Date Solved', 'Email Date'] and value:
                cell.number_format = 'YYYY-MM-DD'
    
    # Set column widths
    ws.column_dimensions['A'].width = 12   # Problem #
    ws.column_dimensions['B'].width = 12   # Difficulty
    ws.column_dimensions['C'].width = 20   # Company
    ws.column_dimensions['D'].width = 40   # Title
    ws.column_dimensions['E'].width = 60   # Description
    ws.column_dimensions['F'].width = 18   # DS Primary
    ws.column_dimensions['G'].width = 18   # DS Subtypes
    ws.column_dimensions['H'].width = 18   # Algo Primary
    ws.column_dimensions['I'].width = 18   # Algo Subtypes
    ws.column_dimensions['J'].width = 15   # Status
    ws.column_dimensions['K'].width = 15   # Date Attempted
    ws.column_dimensions['L'].width = 15   # Date Solved
    ws.column_dimensions['M'].width = 15   # Email Date
    
    # Freeze the header row and add auto-filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # Save the workbook
    wb.save(OUTPUT_PATH)
    print(f"Excel file saved: {OUTPUT_PATH}")
    return all_problems


def main():
    """Main function."""
    print("=" * 70)
    print("Daily Coding Problems - mbox to Excel Converter")
    print("=" * 70)

    # Check if mbox file exists
    if not os.path.exists(MBOX_PATH):
        print(f"ERROR: mbox file not found at {MBOX_PATH}")
        return

    # Load existing problems from spreadsheet (if exists)
    existing_problems = load_existing_problems()

    # Parse mbox file, skipping existing problems
    new_problems, skipped = parse_mbox(existing_problems)

    if not new_problems and not existing_problems:
        print("No problems found in mbox file!")
        return

    if not new_problems:
        print("No new problems to add. Spreadsheet is up to date.")
        return

    # Create/update Excel file
    all_problems = create_excel(new_problems, existing_problems)
    
    # Print summary
    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"Total problems in spreadsheet: {len(all_problems)}")
    print(f"  - Existing (preserved): {len(existing_problems)}")
    print(f"  - New (added): {len(new_problems)}")

    print(f"\nProblems by difficulty:")
    for difficulty in ['Easy', 'Medium', 'Hard']:
        count = sum(1 for p in all_problems if p['Difficulty'] == difficulty)
        print(f"  {difficulty}: {count}")

    if new_problems:
        print(f"\nNewly added problems:")
        for p in new_problems[:10]:  # Show up to 10 new problems
            ds = f" [{p['Data Structures']}]" if p['Data Structures'] else ""
            print(f"  #{p['Problem #']} - {p['Difficulty']} - {p['Company']}{ds}")
        if len(new_problems) > 10:
            print(f"  ... and {len(new_problems) - 10} more")

    solved = sum(1 for p in all_problems if p['Status'] == 'Solved')
    in_progress = sum(1 for p in all_problems if p['Status'] == 'In Progress')
    print(f"\nProgress: {solved} solved, {in_progress} in progress")
    print("=" * 70)
    print("\nDone! Open the Excel file to continue tracking your progress.")


if __name__ == "__main__":
    main()
